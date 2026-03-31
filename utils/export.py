from __future__ import annotations

import io
import json
from typing import Any


def rows_to_excel_bytes(rows: list[dict[str, Any]], sheet_name: str = "Extractions") -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(rows[0].keys()) if rows else []
    if headers:
        ws.append(headers)
        for row in rows:
            out_row = []
            for key in headers:
                value = row.get(key)
                if key == "attributes" and isinstance(value, dict):
                    out_row.append(json.dumps(value))
                elif isinstance(value, (dict, list)):
                    out_row.append(json.dumps(value))
                else:
                    out_row.append(value)
            ws.append(out_row)

        header_font = openpyxl.styles.Font(bold=True)
        header_fill = openpyxl.styles.PatternFill("solid", fgColor="BDD7EE")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = (
                min(max(len(str(cell.value or "")) for cell in col_cells), 60) + 2
            )

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
