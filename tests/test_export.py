import io
import json

import openpyxl
import pytest

from utils.export import rows_to_excel_bytes


@pytest.fixture
def simple_rows():
    return [
        {"class": "medication", "text": "Lisinopril", "attributes": {"medication_group": "Lisinopril"}},
        {"class": "dosage", "text": "10 mg", "attributes": {}},
    ]


def test_returns_bytes(simple_rows):
    result = rows_to_excel_bytes(simple_rows)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_valid_xlsx(simple_rows):
    result = rows_to_excel_bytes(simple_rows)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "class" in headers
    assert "text" in headers


def test_row_count(simple_rows):
    result = rows_to_excel_bytes(simple_rows)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    assert ws.max_row == 3


def test_attributes_flattened(simple_rows):
    result = rows_to_excel_bytes(simple_rows)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    attr_col = headers.index("attributes") + 1
    cell_val = ws.cell(row=2, column=attr_col).value
    parsed = json.loads(cell_val)
    assert isinstance(parsed, dict)


def test_empty_rows():
    result = rows_to_excel_bytes([])
    assert isinstance(result, bytes)
